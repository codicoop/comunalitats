from coopolis.models import ProjectStage, Project, EmploymentInsertion
from cc_courses.models import Activity, Organizer
from dataexports.models import DataExports
from django.http import HttpResponseNotFound, HttpResponse
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill
from django.db.models import Q
import json
from django.conf import settings


class ExportFunctions:
    """ExportFunctions

    This is the generation of excel-like data (in .xlsx) made to fit
    the official formats required for the justification of the
    subsidies.

    Given that each year it changes, and we might need multiple
    documents, or even each Ateneu might need a specific document
    for subsidies that are not the 'conveni', we created a simple
    system to create different functions, 'register' them in the
    admin, and launch them from there.

    This class holds the functions that generate this .xlsx.

    To use them, call callmethod('function_name')
    """

    def __init__(self):
        self.ignore_errors = False
        self.workbook = None
        self.worksheet = None
        self.subsidy_period = 2019

        self.stages_obj = None

        self.subsidy_period_range = None
        self.row_number = 1
        self.error_message = set()
        self.number_of_activities = 0
        self.number_of_stages = 0
        self.number_of_nouniversitaris = 0
        self.number_of_founded_projects = 0

        self.correlations = dict()
        self.organizers = dict()  # Camp per Ateneu / Cercle
        self.d_organizer = None

        # La majoria d'ateneus volen que hi hagi una sola actuació per un
        # projecte encara que hagi tingut diferents tipus d'acompanyament.
        # CoopCamp (i potser algun altre?) volen separar-ho per itineraris,
        # de manera que hi hagi una actuació per l'itinerari de Nova Creació i
        # una pel de Consolidació.
        # Per defecte ho definim per 1 itinerari.
        self.stages_groups = {
            1: 'nova_creacio',
            2: 'nova_creacio',
            6: 'nova_creacio',
            7: 'nova_creacio',
            8: 'nova_creacio',
            9: 'nova_creacio'  # Era Incubació
        }

    def callmethod(self, name):
        if hasattr(self, name):
            obj = DataExports.objects.get(function_name=name)
            self.ignore_errors = obj.ignore_errors
            self.subsidy_period = obj.subsidy_period
            self.subsidy_period_range = obj.subsidy_period.range
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            return getattr(self, name)()
        else:
            return self.return_404("La funció especificada no existeix")

    def return_document(self, name):
        """ Attention: non-ascii characters in the name will cause
        an encoding error with gunicorn.
        Haven't tried it with a proxy under apache, in theory should
        work."""
        if len(self.error_message) > 0 and self.ignore_errors is False:
            return self.return_404()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'), name=name,
        )
        self.workbook.save(response)
        return response

    def return_404(self, message=""):
        """When the exported data has to fit a specific format, there
        are many cases in which we need to stop the generation and tell
        the user that something needs to be fixed.
        This will show a blank page with the message.
        """
        if message:
            self.error_message.add(message)
        message = "<h1>Error al generar el document</h1>" + " ".join(self.error_message)
        return HttpResponseNotFound(message)

    def get_sessions_obj(self, for_minors=False):
        return Activity.objects.filter(
            Q(date_start__range=self.subsidy_period.range, for_minors=for_minors) &
            (
                Q(cofunded__isnull=True) | (
                    Q(cofunded__isnull=False) & Q(cofunded_ateneu=True)
                )
            )
        )

    def import_correlations(self, file_path):
        try:
            file_object = open(file_path, 'r')
            self.correlations = json.load(file_object)
        except FileNotFoundError:
            print(file_path + " not found. ")

    def get_correlation(self, correlated_field, original_data, subsidy_period=2019):
        """When exporting data, we might need to make the exported data
         fit specific requirements. For example, we store the field
         'axis' as 'A', 'B', but the strings we actually need to
         show are:
         'A) Diagnosi i visibilització', 'B) Creació i desenvolupament'

         We have these correlations in a json file and loaded at self.correlations.

        This function is a wrapper to get those.
        """
        try:
            new_data = self.correlations[correlated_field][original_data]
        except KeyError:
            self.error_message.add(
                "<p>El document no s'ha pogut generar perquè s'ha intentat aplicar aquesta correlació:</p"
                "<ul><li>Convocatòria: {}</li><li>Camp: {}</li><li>Dada original: {}</li></ul>"
                "<p>Però no s'ha trobat.</p>".format(subsidy_period, correlated_field, original_data))
            return None
        return new_data

    def create_columns(self, columns):
        """ create_columns

        Expects an iterable containing tuples with the name and the
        width of each column, like this:
        columns = [
            ("First", 40),
            ("Second", 70),
        ]
        """
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            column_letter = get_column_letter(col_num)
            column_dimensions = self.worksheet.column_dimensions[column_letter]
            column_dimensions.font = Font(name="ttf-opensans", size=9)
            column_dimensions.width = column_width
            cell.font = Font(bold=True, name="ttf-opensans", size=9)
            cell.border = Border(bottom=Side(border_style="thin", color='000000'))
            cell.value = str(column_title)

    def fill_row_data(self, row):
        """ fill_row_data

        Populates the columns of a given row with each of the values.
        Expects an iterable with the data for each row:
        row = [
            "first value",
            "second value",
        ]

        Optionally, values can be a tuple to mark the cell as error.
        That will fill the cell with red.
        row = [
            "first value",
            ("second value", True),
        ]
        """
        for col_num, cell_value in enumerate(row, 1):
            cell = self.worksheet.cell(row=self.row_number, column=col_num)
            if isinstance(cell_value, tuple):
                error_mark = cell_value[1]
                if error_mark:
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                cell_value = cell_value[0]
            cell.value = cell_value if isinstance(cell_value, int) else str(cell_value)

    def import_organizers(self):
        # Not forcing any order because we want it in the same order that
        # they see (which should be by ID)
        orgs = Organizer.objects.all()
        if not orgs:
            self.organizers.update({
                0: 'Ateneu'
            })
        else:
            i = 0
            for org in orgs:
                if i == 0:
                    cercle = 'Ateneu'
                else:
                    cercle = f"Cercle {i}"
                self.organizers.update({
                    org.id: cercle
                })
                i += 1
        self.d_organizer = list(self.organizers.keys())[0]

    def get_organizer(self, organizer):
        if not organizer:
            return self.organizers[self.d_organizer]
        return self.organizers[organizer.id]

    """

    Exportació de les memòries d'acompanyament a fitxer de text

    """
    def export_stages_descriptions(self):
        qs = ProjectStage.objects.filter(
            Q(subsidy_period=self.subsidy_period)
            and
            (
                (Q(cofunded__isnull=True))
                or
                (Q(cofunded__isnull=False) and Q(cofunded_ateneu=True))
            )
            and
            Q(follow_up__isnull=False)
        )
        lines = []
        for stage in qs:
            if stage.follow_up != '':
                lines.append(self._html_title(stage.project.name))
                lines.append(self._html_paragraph(stage.follow_up))
        return HttpResponse(self._compose_html(lines))

    @staticmethod
    def _compose_html(lines):
        html = "\r".join(lines)
        html = (
            "<em>Recorda! Fes ctrl+a o cmd+a per seleccionar-ho tot!</em>"
            + html
        )
        html = f"<body style=\"width: 800px\">{html}</body>"
        return html

    @staticmethod
    def _html_title(text):
        return f"<h1>{text}</h1>"

    @staticmethod
    def _html_paragraph(text):
        return f"<p>{text}</p>"

    """
    
    Exportació cofinançades
    
    """
    def export_cofunded_2019_2020(self):
        self.import_correlations(settings.BASE_DIR + "/../apps/dataexports/fixtures/correlations_2019.json")
        """ Each function here called handles the creation of one of the worksheets."""
        self.export_cofunded_actuacions()
        self.export_cofunded_participants()

        return self.return_document("cofinançades")

    def export_cofunded_actuacions(self):
        # Tutorial: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        # Docs: https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook
        self.worksheet.title = "Actuacions"

        columns = [
            ("Eix", 40),
            ("Tipus d'actuació", 70),
            ("Nom de l'actuació", 70),
            ("Data inici d'actuació", 16),
            ("Municipi", 30),
            ("Nombre de participants", 20),
            ("Material de difusió (S/N)", 21),
            ("Incidències", 20),
            ("Cofinançat", 20),
            ("Cofin. Ateneu", 20),
            ("Línia estratègica", 60)
        ]
        self.create_columns(columns)
        self.export_cofunded_actuacions_rows()

    def export_cofunded_actuacions_rows(self):
        obj = Activity.objects.filter(
            date_start__range=self.subsidy_period.range,
            cofunded__isnull=False
        )
        self.number_of_activities = len(obj)
        self.row_number = 1
        for item in obj:
            print(item)
            self.row_number += 1

            axis = self.get_correlation("axis", item.axis)
            if axis is None:
                axis = ("", True)
            subaxis = self.get_correlation("subaxis", item.subaxis)
            if subaxis is None:
                subaxis = ("", True)
            town = None
            if item.place is not None:
                town = item.place.town
            if town is None or town == "":
                town = ("", True)
            material_difusio = "No"
            if item.file1.name:
                material_difusio = "Sí"


            row = [
                axis,
                subaxis,
                item.name,
                item.date_start,
                town,
                item.enrolled.count(),
                material_difusio,
                "",
                item.cofunded,
                "Sí" if item.cofunded_ateneu else "No",
                item.strategic_line if item.strategic_line else ""
            ]
            self.fill_row_data(row)

    def export_cofunded_participants(self):
        self.worksheet = self.workbook.create_sheet("Participants")
        self.row_number = 1

        columns = [
            ("Referència", 40),
            ("Nom actuació", 40),
            ("Cognoms", 20),
            ("Nom", 10),
            ("Doc. identificatiu", 12),
            ("Gènere", 10),
            ("Data naixement", 10),
            ("Municipi del participant", 20),
            ("[Situació laboral]", 20),
            ("[Procedència]", 20),
            ("[Nivell d'estudis]", 20),
            ("[Com ens has conegut]", 20),
            ("[Organitzadora]", 30),
            ("[Email]", 30),
            ("[Telèfon]", 30),
            ("[Projecte]", 30),
            ("[Acompanyaments]", 30),
        ]
        self.create_columns(columns)

        self.export_cofunded_participants_rows()

    def export_cofunded_participants_rows(self):
        activity_reference_number = 0
        obj = Activity.objects.filter(date_start__range=self.subsidy_period.range, cofunded__isnull=False)
        for activity in obj:
            activity_reference_number += 1  # We know that activities where generated first, so it starts at 1.
            for enrollment in activity.confirmed_enrollments:
                participant = enrollment.user
                self.row_number += 1
                if participant.gender is None:
                    gender = ""
                else:
                    gender = self.get_correlation('gender', participant.gender)
                if participant.town is None:
                    town = ""
                else:
                    town = participant.town.name

                row = [
                    f"{activity_reference_number} {activity.name}",  # Referència.
                    activity.name,  # Nom de l'actuació. Camp automàtic de l'excel.
                    participant.surname if participant.surname else "",
                    participant.first_name,
                    participant.id_number,
                    gender if gender else "",
                    participant.birthdate if participant.birthdate else "",
                    town if town else "",
                    participant.get_employment_situation_display() if participant.get_employment_situation_display() else "",
                    participant.get_birth_place_display() if participant.get_birth_place_display() else "",
                    participant.get_educational_level_display() if participant.get_educational_level_display() else "",
                    participant.get_discovered_us_display() if participant.get_discovered_us_display() else "",
                    activity.organizer if activity.organizer else "",
                    participant.email,
                    participant.phone_number if participant.phone_number else "",
                    participant.project if participant.project else "",
                    participant.project.stages_list if participant.project and participant.project.stages_list else ""
                ]
                self.fill_row_data(row)

    """
    
    Exportació Ateneu
    
    """
    def export_2018_2019(self):
        self.import_correlations(settings.BASE_DIR + "/../apps/dataexports/fixtures/correlations_2019.json")
        self.subsidy_period_range = ["2018-11-01", "2019-10-31"]

        """ Each function here called handles the creation of one of the worksheets."""
        self.export_actuacions_2018_2019()
        self.export_stages_2018_2019()
        self.export_founded_projects_2018_2019()
        self.export_participants_2018_2019()
        self.export_nouniversitaris_2018_2019()
        self.export_insercionslaborals_2018_2019()
        self.export_all_projects_2018_2019()

        return self.return_document("justificacio2018-2019")

    def export_2019_2020_dos_itineraris(self):
        self.stages_groups = {
            1: 'nova_creacio',
            2: 'nova_creacio',
            6: 'nova_creacio',
            7: 'consolidacio',
            8: 'consolidacio',
            9: 'incubacio'
        }
        return self.export_2019_2020()

    def export_2019_2020(self):
        self.import_correlations(settings.BASE_DIR + "/../apps/dataexports/fixtures/correlations_2019.json")
        self.subsidy_period_range = ["2019-11-01", "2020-10-31"]

        self.import_organizers()

        """ Each function here called handles the creation of one of the worksheets."""
        self.export_actuacions_2018_2019()
        self.export_stages_2018_2019()
        self.export_founded_projects_2018_2019()
        self.export_participants_2018_2019()
        self.export_nouniversitaris_2018_2019()
        self.export_insercionslaborals_2018_2019()

        return self.return_document("justificacio2019-2020")

    def export_2020_2021_dos_itineraris(self):
        self.stages_groups = {
            1: 'nova_creacio',
            2: 'nova_creacio',
            6: 'nova_creacio',
            7: 'consolidacio',
            8: 'consolidacio',
            9: 'incubacio'
        }
        return self.export_2020_2021()

    def export_2020_2021(self):
        self.import_correlations(settings.BASE_DIR + "/../apps/dataexports/fixtures/correlations_2019.json")

        self.import_organizers()

        """ Each function here called handles the creation of one of the worksheets."""
        self.export_actuacions_2018_2019()
        self.export_stages_2018_2019()
        self.export_founded_projects_2018_2019()
        self.export_participants_2018_2019()
        self.export_nouniversitaris_2018_2019()
        self.export_insercionslaborals_2018_2019()

        return self.return_document("justificacio2019-2020")

    def export_actuacions_2018_2019(self):
        # Tutorial: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        # Docs: https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook
        self.worksheet.title = "Actuacions"

        columns = [
            ("Eix", 40),
            ("Tipus d'actuació", 70),
            ("Nom de l'actuació", 70),
            ("Data inici d'actuació", 16),
            ("Cercle / Ateneu", 16),
            ("Municipi", 30),
            ("Nombre de participants", 20),
            ("Material de difusió (S/N)", 21),
            ("Incidències", 20),
            ("[Entitat]", 20),
            ("[Organitzadora]", 20),
            ("[Lloc]", 20),
            ("[Acció]", 20),
        ]
        self.create_columns(columns)
        self.actuacions_2018_2019_rows_activities()
        self.actuacions_2018_2019_rows_stages()
        self.actuacions_2018_2019_rows_nouniversitaris()
        self.actuacions_2018_2019_rows_founded_projects()
        # Total Stages: self.row_number-Total Activities-1

    def actuacions_2018_2019_rows_activities(self):
        obj = self.get_sessions_obj()
        self.number_of_activities = len(obj)
        for item in obj:
            self.row_number += 1

            axis = self.get_correlation("axis", item.axis)
            if axis is None:
                axis = ("", True)
            subaxis = self.get_correlation("subaxis", item.subaxis)
            if subaxis is None:
                subaxis = ("", True)
            town = None
            if item.place is not None:
                town = item.place.town
            if town is None or town == "":
                town = ("", True)
            material_difusio = "No"
            if item.file1.name:
                material_difusio = "Sí"

            row = [
                axis,
                subaxis,
                item.name,
                item.date_start,
                self.get_organizer(item.organizer),
                town,
                item.enrolled.count(),
                material_difusio,
                "",
                item.entity if item.entity else '',  # Entitat
                item.organizer if item.organizer else '',  # Organitzadora
                item.place if item.place else '',  # Lloc
                item.course,  # Acció
            ]
            self.fill_row_data(row)

    def actuacions_2018_2019_rows_stages(self):
        """
        Acompanyaments que han d'aparèixer:
        - En cas que tingui algun acompanyament de tipus Nova Creació,
          ha d'aparèixer fent la suma d'hores de tots els acompanyaments
          d'aquet tipus.
        - Idem però pels acompanyaments de Consolidació.
        - Idem però pels acompanyaments d'Incubació.
        Per tant com a màxim apareixerà 3 vegades.

        TIPOLOGIA COVID: Pendent de saber què n'he de fer.
        """
        obj = ProjectStage.objects.order_by('date_start').filter(
            subsidy_period=self.subsidy_period
        )
        self.stages_obj = {}
        for item in obj:
            if int(item.stage_type) not in self.stages_groups:
                continue
            group = self.stages_groups[int(item.stage_type)]
            p_id = item.project.id
            if p_id not in self.stages_obj:
                self.stages_obj.update({
                    p_id: {}
                })
            if group not in self.stages_obj[p_id]:
                self.stages_obj[p_id].update({
                    group: {
                        'obj': item,
                        'total_hours': 0,
                        'participants': []
                    }
                })
            if item.hours is None:
                item.hours = 0
            self.stages_obj[p_id][group]['total_hours'] += item.hours

            # Aprofitem per omplir les dades dels participants aquí per no
            # repetir el procés més endavant. La qüestió és que encara que un
            # participant hagi participat a diversos acompanyaments, aquí
            # només aparegui una vegada.
            for participant in item.involved_partners.all():
                if (participant
                        not in self.stages_obj[p_id][group]['participants']):
                    self.stages_obj[p_id][group]['participants'].append(
                        participant
                    )
        """
        En aquest punt, self.stages_obj té aquesta estructura:
        160: {
            'nova_creacio': {
                'obj': '<ProjectStage: Consell Comarcal Conca de Barberà - Concactiva: 00 Nova creació - acollida>',
                'total_hours': 3,
                'participants': [
                         '<User: Jordi París>', '<User: Francesc Viñas>',
                         '<User: Carme Pallàs>',
                         '<User: Pere Picornell Busquets>'
                ],
                'row_number': 122
            }
        },
        20: {
            'consolidacio': {
                'obj': '<ProjectStage: La Providència SCCL: 04 Consolidació - acompanyament>',
                'total_hours': 36,
                'participants': [
                    '<User: Teresa Trilla Ferré>',
                    '<User: Marc Trilla Güell>',
                    '<User: Gerard Nogués Balsells>',
                    '<User: tais bastida aubareda>'
                ],
                'row_number': 126},
            'nova_creacio': {
                'obj': '<ProjectStage: La Providència SCCL: 02 Nova creació - constitució>',
                'total_hours': 7,
                'participants': [
                    '<User: Marc Trilla Güell>',
                    '<User: Esther Perello Piulats>'
                ],
                'row_number': 127
            }
        }, 
        
        És a dir, cada Projecte té un element amb la seva ID, que conté els
        grups que corresponguin.
        Cada grup conté:
            - Una instància de l'objecte ProjectStage (el primer que hagi pillat)
            - 'total_hours', amb la suma de totes les hores de tots els ProjectStages
                del mateix grup i projecte.
        """

        """
        A continuació el que farem és afegir com a row cada un d'aquests grups.
        De manera que cada itinerari (creació, consolidació..) de cada projecte
        tindrà la seva propia fila a Actuacions.
        
        En el procés de fer-ho, aprofitem per desar el nº de row dins 
        self.stages_obj[id_projecte][nom_grup][row_number]
        per poder-ho fer servir més endavant quan generem els rows de les 
        persones participants.
        """

        self.number_of_stages = 0
        for project_id, project in self.stages_obj.items():
            for group_name, group in project.items():
                self.number_of_stages += 1
                item = group['obj']
                self.row_number += 1

                # Desant el nº per quan fem el llistat de participants. El
                # self.row_number conté el row REAL, que com que inclou la fila
                # de headers, és un més que la que s'assignarà com a
                # referència.
                self.stages_obj[project_id][group_name][
                    'row_number'
                ] = self.row_number - 1

                axis = self.get_correlation("axis", item.axis)
                if axis is None:
                    axis = ("", True)
                subaxis = self.get_correlation("subaxis", item.subaxis)
                if subaxis is None:
                    subaxis = ("", True)
                town = item.project.town
                if town is None or town == "":
                    town = ("", True)

                row = [
                    axis,
                    subaxis,
                    item.project.name,
                    item.date_start if not None else '',
                    self.get_organizer(item.stage_organizer),
                    town,
                    len(group['participants']),  # Nombre de participants
                    "No",
                    "",
                    item.entity if item.entity else '',  # Entitat
                    item.stage_organizer if item.stage_organizer else '',  # Organitzadora
                    '(no aplicable)',  # Lloc
                    '(no aplicable)',  # Acció
                ]
                self.fill_row_data(row)

    def actuacions_2018_2019_rows_nouniversitaris(self):
        obj = self.get_sessions_obj(for_minors=True)
        self.number_of_nouniversitaris = len(obj)
        for item in obj:
            self.row_number += 1

            axis = self.get_correlation("axis", item.axis)
            if axis is None:
                axis = ("", True)
            subaxis = self.get_correlation("subaxis", item.subaxis)
            if subaxis is None:
                subaxis = ("", True)
            town = None
            if item.place is not None:
                town = item.place.town
            if town is None or town == "":
                town = ("", True)

            row = [
                axis,
                subaxis,
                item.name,
                item.date_start,
                self.get_organizer(item.organizer),
                town,
                item.minors_participants_number,
                "No",
                "",
                item.entity if item.entity else '',  # Entitat
                item.organizer if item.organizer else '',  # Organitzadora
                item.place if item.place else '',  # Lloc
                item.course,  # Acció
            ]
            self.fill_row_data(row)

    def actuacions_2018_2019_rows_founded_projects(self):
        """
        Tots els projectes que tinguin data de constitució dins de les dates de la convocatòria apareixeran
        a la pestanya d'EntitatsCreades.
        No obstant només aquells que tinguin una actuació vinculada durant el període de la convocatòria han de
        d'aparèixer a la pestanya d'Actuacions.

        Els que tenen això vol dir que hi ha hagut un acompanyament del projecte dins de la convocatòria.
        Poden haver-hi hagut varis acompanyaments, per tant, hem de d'obtenir l'acompanyament més recent.

        Si tot això existeix mostrem les dades del més recent, sinó, ignorem el projecte.

        Després a la pestanya d'EntitatsCreades hem de fer el mateix filtre per saber quins tenen
        actuació creada, i deduïr per l'ordre quina ID li toca.
        """
        obj = Project.objects.filter(constitution_date__range=self.subsidy_period_range)
        self.number_of_founded_projects = len(obj)
        for project in obj:
            stages = ProjectStage.objects.filter(
                project=project, subsidy_period=self.subsidy_period
            ).order_by("-date_start")[:1]
            if stages.count() < 1:
                continue
            stage = stages.all()[0]

            self.row_number += 1
            axis = self.get_correlation("axis", stage.axis)
            if axis is None:
                axis = ("", True)
            subaxis = self.get_correlation("subaxis", stage.subaxis)
            if subaxis is None:
                subaxis = ("", True)
            town = project.town
            if town is None or town == "":
                town = ("", True)

            row = [
                axis,
                subaxis,
                project.name,
                stage.date_start,
                self.get_organizer(stage.stage_organizer),
                town,
                stage.involved_partners.count(),
                "No",
                "",
                stage.entity if stage.entity else '',  # Entitat
                stage.stage_organizer if stage.stage_organizer else '',  # Organitzadora
                '(no aplicable)',  # Lloc
                '(no aplicable)',  # Acció
            ]
            self.fill_row_data(row)

    def export_stages_2018_2019(self):
        self.worksheet = self.workbook.create_sheet("Acompanyaments")
        self.row_number = 1

        columns = [
            ("Referència", 20),
            ("Nom actuació", 40),
            ("Destinatari de l'acompanyament (revisar)", 45),
            ("En cas d'entitat (nom de l'entitat)", 40),
            ("En cas d'entitat (revisar)", 30),
            ("Creació/consolidació", 18),
            ("Data d'inici", 13),
            ("Localitat", 20),
            ("Breu descripció del projecte", 50),
            ("Total hores d'acompanyament", 10),
            ("[Data fi]", 13),
        ]
        self.create_columns(columns)

        self.stages_2018_2019_rows()

    def stages_2018_2019_rows(self):
        reference_number = self.number_of_activities

        for p_id, stage in self.stages_obj.items():
            for group_name, group in stage.items():
                self.row_number += 1
                reference_number += 1
                item = group['obj']

                # hours = item.hours if item.hours is not None else ("", True)
                hours = group['total_hours']
                town = item.project.town if item.project.town is not None else ("", True)
                crea_consolida = self.get_correlation("stage_type", item.stage_type)
                row = [
                    f"{reference_number} {item.project.name}",  # Referència.
                    item.project.name,  # Camp no editable, l'ha d'omplir l'excel automàticament.
                    "Entitat",
                    # "Destinatari de l'actuació" Opcions: Persona física/Promotor del projecte/Entitat PENDENT.
                    item.project.name,  # "En cas d'entitat (Nom de l'entitat)"
                    self.get_correlation("project_status", item.project.project_status),
                    crea_consolida if crea_consolida else '',  # "Creació/consolidació".
                    item.date_start if item.date_start else '',
                    town,
                    item.project.description,  # Breu descripció.
                    hours,  # Total hores d'acompanyament.
                    item.date_end if item.date_end else '',
                ]
                self.fill_row_data(row)

    def export_founded_projects_2018_2019(self):
        self.worksheet = self.workbook.create_sheet("EntitatCreada")
        self.row_number = 1

        columns = [
            ("Referència", 10),
            ("Nom actuació", 40),
            ("Nom de l'entitat", 40),
            ("NIF de l'entitat", 12),
            ("Nom i cognoms persona de contacte", 30),
            ("Correu electrònic", 12),
            ("Telèfon", 10),
            ("Economia solidària (revisar)", 35),
            ("Ateneu / Cercle (omplir a ma)", 35),
            ("[Acompanyaments]", 10),
        ]
        self.create_columns(columns)

        self.founded_projects_2018_2019_rows()

    def founded_projects_2018_2019_rows(self):
        # The Ids start at 1, so later we add 1 to this number to have the right ID.
        founded_projects_reference_number = \
            self.number_of_stages + self.number_of_activities + self.number_of_nouniversitaris
        obj = Project.objects.filter(constitution_date__range=self.subsidy_period_range)
        for project in obj:
            # Repeating the same filter than in Actuacions to determine if we have an Actuació or not
            reference_number = ""
            name = ""
            stages = ProjectStage.objects.filter(
                project=project, subsidy_period=self.subsidy_period
            ).order_by("-date_start")[:1]
            if stages.count() > 0:
                stage = stages.all()[0]
                founded_projects_reference_number += 1
                reference_number = f"{founded_projects_reference_number} {project.name}"
                name = project.name

            self.row_number += 1
            if project.cif is None:
                self.error_message.add(
                    "<p><strong>Error: falta NIF</strong>. L'entitat '{}' apareix com a EntitatCreada"
                    " perquè té una Data de constitució dins de la convocatòria, però si no té NIF, "
                    "no pot ser inclosa a l'excel.</p>".format(project.name))
                project.cif = ""
            row = [
                reference_number,  # Referència. En aquest full no cal que tinguin relació amb Actuacions.
                name,  # Nom de l'actuació. En aquest full no cal que tinguin relació amb Actuacions.
                project.name,
                project.cif,
                project.partners.all()[0].full_name if project.partners.all() else "",
                project.mail,
                project.phone,
                "Sí",  # Economia solidària
                "",  # Ateneu / Cercle
                project.stages_list
            ]
            self.fill_row_data(row)

    def export_participants_2018_2019(self):
        self.worksheet = self.workbook.create_sheet("Participants")
        self.row_number = 1

        columns = [
            ("Referència", 40),
            ("Nom actuació", 40),
            ("Cognoms", 20),
            ("Nom", 10),
            ("Doc. identificatiu", 12),
            ("Gènere", 10),
            ("Data naixement", 10),
            ("Municipi del participant", 20),
            ("[Situació laboral]", 20),
            ("[Procedència]", 20),
            ("[Nivell d'estudis]", 20),
            ("[Com ens has conegut]", 20),
            ("[Organitzadora]", 30),
            ("[Email]", 30),
            ("[Telèfon]", 30),
            ("[Projecte]", 30),
            ("[Acompanyaments]", 30),
        ]
        self.create_columns(columns)

        self.participants_2018_2019_rows()
        self.participants_project_stages_rows()

    def participants_project_stages_rows(self):
        for project_id, project in self.stages_obj.items():
            for group_name, group in project.items():
                activity = group['obj']
                activity_reference_number = group['row_number']
                for participant in group['participants']:
                    if participant.gender is None:
                        gender = ""
                    else:
                        gender = self.get_correlation('gender', participant.gender)
                    if participant.town is None:
                        town = ""
                    else:
                        town = participant.town.name

                    row = [
                        f"{activity_reference_number} {activity.project.name}",  # Referència.
                        activity.project.name,  # Nom de l'actuació. Camp automàtic de l'excel.
                        participant.surname or "",
                        participant.first_name,
                        participant.id_number,
                        gender if gender else "",
                        participant.birthdate or "",
                        town if town else "",
                        participant.get_employment_situation_display() or "",
                        participant.get_birth_place_display() or "",
                        participant.get_educational_level_display() or "",
                        participant.get_discovered_us_display() or "",
                        activity.stage_organizer or "",
                        participant.email,
                        participant.phone_number or "",
                        participant.project or "",
                        participant.project.stages_list if participant.project and participant.project.stages_list else "",
                    ]
                    self.row_number += 1
                    self.fill_row_data(row)

    def participants_2018_2019_rows(self):
        activity_reference_number = 0
        obj = self.get_sessions_obj(for_minors=False)
        for activity in obj:
            activity_reference_number += 1  # We know that activities where generated first, so it starts at 1.
            for enrollment in activity.confirmed_enrollments:
                participant = enrollment.user
                self.row_number += 1
                if participant.gender is None:
                    gender = ""
                else:
                    gender = self.get_correlation('gender', participant.gender)
                if participant.town is None:
                    town = ""
                else:
                    town = participant.town.name

                row = [
                    f"{activity_reference_number} {activity.name}",  # Referència.
                    activity.name,  # Nom de l'actuació. Camp automàtic de l'excel.
                    participant.surname if participant.surname else "",
                    participant.first_name,
                    participant.id_number,
                    gender if gender else "",
                    participant.birthdate if participant.birthdate else "",
                    town if town else "",
                    participant.get_employment_situation_display() or "",
                    participant.get_birth_place_display() or "",
                    participant.get_educational_level_display() or "",
                    participant.get_discovered_us_display() or "",
                    activity.organizer if activity.organizer else "",
                    participant.email,
                    participant.phone_number or "",
                    participant.project if participant.project else "",
                    participant.project.stages_list if participant.project and participant.project.stages_list else "",
                ]
                self.fill_row_data(row)

    def export_nouniversitaris_2018_2019(self):
        self.worksheet = self.workbook.create_sheet("ParticipantsNoUniversitaris")
        self.row_number = 1

        columns = [
            ("Referència", 40),
            ("Nom actuació", 40),
            ("Grau d'estudis", 20),
            ("Nom centre educatiu", 20),
        ]
        self.create_columns(columns)

        self.nouniversitaris_2018_2019_rows()

    def nouniversitaris_2018_2019_rows(self):
        nouniversitari_reference_number = self.number_of_stages + self.number_of_activities
        obj = self.get_sessions_obj(for_minors=True)
        for activity in obj:
            self.row_number += 1
            nouniversitari_reference_number += 1
            row = [
                f"{nouniversitari_reference_number} {activity.name}",  # Referència.
                activity.name,  # Nom de l'actuació. Camp automàtic de l'excel.
                self.get_correlation('minors_grade', activity.minors_grade),
                activity.minors_school_name,
            ]
            self.fill_row_data(row)

    def export_insercionslaborals_2018_2019(self):
        self.worksheet = self.workbook.create_sheet("InsercionsLaborals")
        self.row_number = 1

        columns = [
            ("Referència (omplir a ma)", 20),
            ("Nom actuació", 20),
            ("Cognoms", 20),
            ("Nom", 20),
            ("ID", 20),
            ("Data alta", 20),
            ("Data baixa", 20),
            ("Tipus contracte", 20),
            ("Gènere", 20),
            ("Data naixement", 20),
            ("Població", 20),
            ("NIF Projecte", 20),
            ("Nom projecte", 20),
            ("Cercle / Ateneu (omplir a ma)", 20),
            ("[ convocatòria ]", 20),
        ]
        self.create_columns(columns)

        self.insercionslaborals_2018_2019_rows()

    def insercionslaborals_2018_2019_rows(self):
        obj = EmploymentInsertion.objects.filter(
            subsidy_period__date_start__range=self.subsidy_period_range)
        for insertion in obj:
            self.row_number += 1
            id_number = insertion.user.id_number
            if not id_number:
                id_number = ('', True)
            insertion_date = insertion.insertion_date
            if not insertion_date:
                insertion_date = ('', True)
            contract_type = insertion.get_contract_type_display()
            if not contract_type:
                contract_type = ('', True)
            birthdate = insertion.user.birthdate
            if not birthdate:
                birthdate = ('', True)
            town = insertion.user.town
            if not town:
                town = ('', True)
            cif = insertion.project.cif
            if not cif:
                cif = ('', True)

            if insertion.user.gender is None:
                gender = ""
            else:
                gender = self.get_correlation('gender', insertion.user.gender)

            row = [
                '',  # Deixem referència en blanc pq la posin a ma.
                '',  # Nom actuació
                insertion.user.surname,
                insertion.user.first_name,  # Persona
                id_number,
                insertion_date,  # Data d'alta SS
                '',  # Data baixa SS
                contract_type,  # Tipus de contracte
                gender,
                birthdate,
                town,
                cif,
                insertion.project.name,  # Projecte
                insertion.subsidy_period,  # Convocatòria
                '',  # Cercle / Ateneu
            ]
            self.fill_row_data(row)

    def export_all_projects_2018_2019(self):
        self.worksheet = self.workbook.create_sheet("PROJECTES")
        self.row_number = 1

        columns = [
            ("ID", 5),
            ("Data registre", 12),
            ("Data constitució", 12),
            ("Nom de l'entitat", 40),
            ("NIF de l'entitat", 12),
            ("Nom i cognoms persona de contacte", 30),
            ("Correu electrònic", 30),
            ("Telèfon", 15),
            ("[Acompanyaments]", 40),
            ("[Eixos]", 40),
        ]
        self.create_columns(columns)

        self.all_projects_2018_2019_rows()

    def all_projects_2018_2019_rows(self):
        self.row_number = 1
        obj = Project.objects.order_by('id').all()
        for project in obj:
            self.row_number += 1
            row = [
                project.id,
                project.registration_date if project.registration_date else "",
                project.constitution_date if project.constitution_date else "",
                project.name,
                project.cif if project.cif else "",
                project.partners.all()[0].full_name if project.partners.all() else "",
                project.mail,
                project.phone,
                project.stages_list if project.stages_list else "",
                project.axis_list if project.axis_list else ""
            ]
            self.fill_row_data(row)
