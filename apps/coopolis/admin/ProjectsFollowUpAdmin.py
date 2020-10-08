from datetime import datetime
from urllib.parse import urlencode

from django.contrib import admin
from django.db.models import Count, Q, Sum, OuterRef, Subquery, IntegerField
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import path, reverse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .ProjectAdmin import FilterByFounded
from dataexports.models import SubsidyPeriod
from coopolis.models.projects import ProjectStage
from ..models import User


class MissingCurrentSubsidyPeriod(Exception):
    pass


class ProjectsFollowUpAdmin(admin.ModelAdmin):
    """
    Inspired in: https://medium.com/@hakibenita/how-to-turn-django-admin-into-a-lightweight-dashboard-a0e0bbf609ad
    """
    class Media:
        js = ('js/grappellihacks.js',)
        css = {
            'all': ('styles/grappellihacks.css',)
        }

    change_list_template = 'admin/projects_follow_up.html'
    list_filter = (
        'subsidy_period', 'follow_up_situation',
        FilterByFounded,
        ('stages__stage_responsible', admin.RelatedOnlyFieldListFilter),
        'project_status',
    )
    search_fields = ('id', 'name', 'cif', )
    show_full_result_count = False
    list_display = ('name', )
    list_per_page = 99999

    def get_rows(self, projects, request):
        ctxt = {}
        project_ids = {}
        for project in projects:
            project_ids.update({project.id: project})

        filtered_subsidy_period = None
        if 'subsidy_period' in request.GET:
            filtered_subsidy_period = request.GET['subsidy_period']
        else:
            current = SubsidyPeriod.get_current()
            if current:
                filtered_subsidy_period = current.id

        query = {
            # We need this count just to force the group_by
            'grouping': Count('project_id'),
            'members_h': Subquery(
                User.objects
                    .filter(gender='MALE')
                    .get_num_members_for_project(OuterRef('project_id')),
                output_field=IntegerField()
            ),
            'members_d': Subquery(
                User.objects
                    .filter(gender='FEMALE')
                    .get_num_members_for_project(OuterRef('project_id')),
                output_field=IntegerField()
            ),
            'members_total': Subquery(
                User.objects
                    .filter()
                    .get_num_members_for_project(OuterRef('project_id')),
                output_field=IntegerField()
            ),
            'acollida_hores': Sum(
                'hours',
                filter=Q(stage_type=1)
            ),
            'acollida_certificat':
                Count(
                    'scanned_certificate',
                    filter=(
                        Q(stage_type=1) &
                        Q(scanned_certificate__isnull=False) &
                        ~Q(scanned_certificate__exact='')
                    )
                ),
            'proces_hores': Sum(
                'hours',
                filter=Q(stage_type=2)
            ),
            'proces_certificat':
                Count(
                    'scanned_certificate',
                    filter=(
                        Q(stage_type=2) &
                        Q(scanned_certificate__isnull=False) &
                        ~Q(scanned_certificate__exact='')
                    )
                ),
            'constitucio_hores': Sum(
                'hours',
                filter=Q(stage_type=6)
            ),
            'constitucio_certificat':
                Count(
                    'scanned_certificate',
                    filter=(
                        Q(stage_type=6) &
                        Q(scanned_certificate__isnull=False) &
                        ~Q(scanned_certificate__exact='')
                    )
                ),
            'consolidacio_hores': Sum(
                'hours',
                filter=Q(stage_type__in=[7, 8])
            ),
            'consolidacio_certificat':
                Count(
                    'scanned_certificate',
                    filter=(
                        Q(stage_type__in=[7, 8]) &
                        Q(scanned_certificate__isnull=False) &
                        ~Q(scanned_certificate__exact='')
                    )
                ),
        }

        qs_project_stages = ProjectStage.objects.filter(
            project_id__in=project_ids
        )
        if filtered_subsidy_period:
            qs_project_stages = qs_project_stages.filter(
                subsidy_period=filtered_subsidy_period
            )

        # Annotate adds columns to each row with the sum or calculations of
        qs_project_stages = qs_project_stages.values('project_id').annotate(**query)
        # Order_by fucks up the group by
        qs_project_stages = qs_project_stages.order_by()

        ctxt['rows'] = qs_project_stages

        # Normally it should be easier to call aggregate to have the totals,
        # but given how complex is it to combine
        # with the ORM filters, I opted for filling the values like this.
        # The original version was: qs.aggregate(**query)
        totals = dict(
            total_members_h=0,
            total_members_d=0,
            total_members_total=0,
            total_acollida_hores=0,
            total_acollida_certificat=0,
            total_proces_hores=0,
            total_proces_certificat=0,
            total_constitucio_hores=0,
            total_constitucio_certificat=0,
            total_consolidacio_hores=0,
            total_consolidacio_certificat=0,
            total_employment_insertions=0,
            total_constitutions=0,
        )
        for row in ctxt['rows']:
            row['project'] = project_ids[row['project_id']]
            print(row['project'].follow_up_with_date)
            totals['total_members_h'] += (
                row['members_h']if row['members_h']else 0
            )
            totals['total_members_d'] += (
                row['members_d']if row['members_d']else 0
            )
            totals['total_members_total'] += (
                row['members_total']if row['members_total']else 0
            )
            totals['total_acollida_hores'] += (
                row['acollida_hores']if row['acollida_hores']else 0
            )
            totals['total_acollida_certificat'] += (
                1 if row['acollida_certificat']else 0
            )
            totals['total_proces_hores'] += (
                row['proces_hores']if row['proces_hores']else 0
            )
            totals['total_proces_certificat'] += (
                1 if row['proces_certificat']else 0
            )
            totals['total_constitucio_hores'] += (
                row['constitucio_hores']if row['constitucio_hores']else 0
            )
            totals['total_constitucio_certificat'] += (
                1 if row['constitucio_certificat']else 0
            )
            totals['total_consolidacio_hores'] += (
                row['consolidacio_hores']if row['consolidacio_hores']else 0
            )
            totals['total_consolidacio_certificat'] += (
                1 if row['consolidacio_certificat']else 0
            )
            totals['total_employment_insertions'] += (
                len(row['project'].employment_insertions.all())
                if row['project'].employment_insertions else 0
            )
            totals['total_constitutions'] += (
                1 if row['project'].constitution_date else 0
            )

        ctxt['totals'] = totals

        return ctxt

    def get_download_url(self, request):
        querystring = urlencode(request.GET)
        url = reverse('admin:followup-spreadsheet')
        url = f"{url}?{querystring}"
        return {'spreadsheet_url': url}

    def changelist_view(self, request, extra_context=None):
        # They usually want to use the current period by default.
        if 'subsidy_period__id__exact' not in request.GET:
            return self.redirect_to_current_period(request)

        response = super().changelist_view(
            request,
            extra_context=extra_context,
        )

        # Getting queryset with filters applied
        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response

        # Els filtres fan joins, necessitem el Count per forçar un group by.
        projects = qs.annotate(Count('id'))

        rows = self.get_rows(projects, request)
        url = self.get_download_url(request)
        response.context_data = {
            **response.context_data,
            **rows,
            **url
        }
        return response

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'spreadsheet/',
                self.admin_site.admin_view(self.download_spreadsheet),
                name='followup-spreadsheet',
            ),
        ]
        return custom_urls + urls

    @staticmethod
    def redirect_to_current_period(request):
        current = SubsidyPeriod.get_current()
        if current:
            value = str(current.id)
        else:
            raise MissingCurrentSubsidyPeriod(
                "El llistat per defecte es carrega mostrant la convocatòria "
                "vigent, però ha fallat a l'intentar-la trobar. Si us plau "
                "revisa que existeixi una convocatòria creada per la data "
                "actual."
            )
        query_string = request.META['QUERY_STRING']
        value = f"subsidy_period__id__exact={value}"
        if query_string != '':
            value = f"&{value}"
        query_string = query_string + value
        return HttpResponseRedirect(f"{request.path_info}?{query_string}")

    def download_spreadsheet(self, request):
        # Getting queryset with filters applied
        try:
            qs = super().get_changelist_instance(request).queryset
        except (AttributeError, KeyError) as e:
            return request

        rows = self.get_rows(qs)
        sheet = FollowUpSpreadsheet(rows['rows'], rows['totals'])
        return sheet.export_seguiment_acompanyaments()


class FollowUpSpreadsheet:
    def __init__(self, raw_rows, raw_totals):
        self.row_number = 1
        self.raw_rows = raw_rows
        self.raw_totals = raw_totals
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def return_document(self, name):
        """ Attention: non-ascii characters in the name will cause
        an encoding error with gunicorn.
        Haven't tried it with a proxy under apache, in theory should
        work."""
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        date = datetime.now().strftime('%Y-%m-%d')
        response['Content-Disposition'] = (
            f'attachment; filename={date}-{name}.xlsx'
        )
        self.workbook.save(response)
        return response

    def create_columns(self, columns):
        """ create_columns

        Expects an iterable containing tuples with the name and the
        width of each column, like this:
        columns = [
            ("First", 40),
            ("Second", 70),
        ]
        """
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            column_letter = get_column_letter(col_num)
            column_dimensions = self.worksheet.column_dimensions[column_letter]
            column_dimensions.font = Font(name="ttf-opensans", size=9)
            column_dimensions.width = column_width
            cell.font = Font(bold=True, name="ttf-opensans", size=9)
            cell.border = Border(bottom=Side(border_style="thin", color='000000'))
            cell.value = str(column_title)

    def fill_row_data(self, row):
        """ fill_row_data

        Populates the columns of a given row with each of the values.
        Expects an iterable with the data for each row:
        row = [
            "first value",
            "second value",
        ]

        Optionally, values can be a tuple to mark the cell as error.
        That will fill the cell with red.
        row = [
            "first value",
            ("second value", True),
        ]
        """
        for col_num, cell_value in enumerate(row, 1):
            cell = self.worksheet.cell(row=self.row_number, column=col_num)
            if isinstance(cell_value, tuple):
                error_mark = cell_value[1]
                if error_mark:
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                cell_value = cell_value[0]
            cell.value = cell_value if isinstance(cell_value, int) else str(cell_value)

    def export_seguiment_acompanyaments(self):
        """ Each function here called handles the creation of one of the worksheets."""
        self.export_seguiment_acompanyaments_sheet1()

        return self.return_document("seguiment_acompanyaments")

    def export_seguiment_acompanyaments_sheet1(self):
        # Tutorial: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        # Docs: https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook
        self.worksheet.title = "Acompanyaments"

        columns = [
            ("ID", 10),
            ("Seguiment", 35),
            ("Organitzadora", 20),
            ("Nom", 50),
            ("Eix-Subeix", 30),
            ("Tutoritza", 20),
            ("Membres H", 15),
            ("Membres D", 15),
            ("Total sòcies", 15),
            ("Sector", 20),
            ("Descripció", 60),
            ("Estat", 20),
            ("Territori", 20),
            ("Acollida H.", 15),
            ("Acollida cert.", 15),
            ("Procés H.", 15),
            ("Procés cert.", 15),
            ("Constitució H.", 15),
            ("Constitució cert.", 15),
            ("Consolidació H.", 15),
            ("Consolidació cert.", 15),
            ("Insercions previstes", 15),
            ("Insercions justificades", 15),
            ("Constitució", 15),
            ("Altres", 60),
        ]
        self.create_columns(columns)
        self.export_seguiment_acompanyaments_rows()

    def export_seguiment_acompanyaments_rows(self):
        self.row_number = 1
        for raw_row in self.raw_rows:
            self.row_number += 1
            stage_organizer = None
            stage_responsible = None
            if len(raw_row.stages.all()) > 0:
                stage_organizer = raw_row.stages.all().last().stage_organizer
                stage_responsible = raw_row.stages.all().last().stage_responsible

            row = [
                raw_row.id,
                raw_row.get_follow_up_situation_display()
                if raw_row.get_follow_up_situation_display() else '',
                stage_organizer if stage_organizer else '',
                raw_row.name,
                raw_row.axis_list if raw_row.axis_list else '',
                stage_responsible if stage_responsible else '',
                raw_row.members_h,
                raw_row.members_d,
                raw_row.members_total,
                raw_row.get_sector_display(),
                raw_row.description if raw_row.description else '',
                raw_row.get_project_status_display(),
                raw_row.full_town_district
                    if raw_row.full_town_district else '',
                raw_row.acollida_hores if raw_row.acollida_hores else 0,
                1 if raw_row.acollida_certificat > 0 else 0,
                raw_row.proces_hores if raw_row.proces_hores else 0,
                1 if raw_row.proces_certificat > 0 else 0,
                raw_row.constitucio_hores if raw_row.constitucio_hores else 0,
                1 if raw_row.constitucio_certificat > 0 else 0,
                raw_row.consolidacio_hores if raw_row.consolidacio_hores else 0,
                1 if raw_row.consolidacio_certificat > 0 else 0,
                raw_row.employment_estimation,
                len(raw_row.employment_insertions.all()),
                raw_row.constitution_date if raw_row.constitution_date else '',
                raw_row.other if raw_row.other else '',
            ]
            self.fill_row_data(row)


class ConstitutionDateFilter(admin.SimpleListFilter):
    title = "Amb data de constitució dins la convocatòria…"

    parameter_name = 'constitution_subsidy'

    def lookups(self, request, model_admin):
        qs = SubsidyPeriod.objects.all()
        qs.order_by('name')
        return list(qs.values_list('id', 'name'))

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            period = SubsidyPeriod.objects.get(id=value)
            return queryset.filter(constitution_date__range=(period.date_start, period.date_end))
        return queryset


class ProjectsConstitutedAdmin(admin.ModelAdmin):
    """
    Inspired in: https://medium.com/@hakibenita/how-to-turn-django-admin-into-a-lightweight-dashboard-a0e0bbf609ad
    """
    class Media:
        js = ('js/grappellihacks.js',)
        css = {
            'all': ('styles/grappellihacks.css',)
        }

    change_list_template = 'admin/projects_constituted.html'
    list_filter = (ConstitutionDateFilter, )
    show_full_result_count = False
    list_display = ('name', )
    list_per_page = 99999

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(
            request,
            extra_context=extra_context,
        )

        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response

        query = {
            'members_h': Count('stages__involved_partners',
                               filter=Q(stages__involved_partners__gender='MALE'), distinct=True),
            'members_d': Count('stages__involved_partners',
                               filter=Q(stages__involved_partners__gender='FEMALE'), distinct=True),
            'members_a': Count('stages__involved_partners',
                               filter=~Q(stages__involved_partners__gender='FEMALE') &
                                      ~Q(stages__involved_partners__gender='MALE'), distinct=True),
        }

        # Annotate adds columns to each row with the sum or calculations of the row:
        response.context_data['rows'] = list(
            qs.filter(cif__isnull=False, constitution_date__isnull=False).annotate(**query)
        )

        # Normally it should be easier to call aggregate to have the totals, but given how complex is it to combine
        # with the ORM filters, I opted for filling the values like this.
        # The original version was: qs.aggregate(**query)
        totals = dict(
            total_members_h=0,
            total_members_d=0,
            total_members_a=0,
        )
        for row in response.context_data['rows']:
            totals['total_members_h'] += row.members_h if row.members_h else 0
            totals['total_members_d'] += row.members_d if row.members_d else 0
            totals['total_members_a'] += row.members_a if row.members_a else 0

        response.context_data['totals'] = totals

        return response

    # def has_change_permission(self, request, obj=None):
    #     return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False
