from datetime import datetime
from urllib.parse import urlencode

from django.contrib import admin
from django.db.models import Count, Q, Sum, OuterRef, Subquery, IntegerField
from django.http import HttpResponse
from django.urls import path, reverse
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from .ProjectAdmin import FilterByFounded
from apps.dataexports.models import SubsidyPeriod
from apps.coopolis.models.projects import ProjectStage, ProjectsFollowUp, \
    ProjectsFollowUpService, ProjectsConstituted, ProjectsConstitutedService
from ..mixins import FilterByCurrentSubsidyPeriodMixin
from ..models import User


@admin.register(ProjectsFollowUp)
class ProjectsFollowUpAdmin(FilterByCurrentSubsidyPeriodMixin, admin.ModelAdmin):
    """
    Deprecated: from Nov 2021 the updated report is ProjectsFollowUpAdmin.
    Keeping the deprecated one until we are sure that it's not needed anymore.

    Inspired in: https://medium.com/@hakibenita/how-to-turn-django-admin-into-a-lightweight-dashboard-a0e0bbf609ad
    """
    class Media:
        js = ('js/grappellihacks.js',)
        css = {
            'all': ('styles/grappellihacks.css',)
        }

    change_list_template = 'admin/projects_follow_up.html'
    list_filter = (
        'stages__subsidy_period', 'follow_up_situation',
        FilterByFounded,
        ('stages__stage_responsible', admin.RelatedOnlyFieldListFilter),
        'project_status',
    )
    search_fields = ('id', 'name', 'cif', )
    show_full_result_count = False
    list_display = ('name', )
    list_per_page = 99999
    subsidy_period_filter_param = 'stages__subsidy_period__id__exact'

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(
            request,
            extra_context=extra_context,
        )

        # Getting queryset with filters applied
        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response

        # Els filtres fan joins, necessitem el Count per forçar un group by.
        projects = qs.annotate(Count('id'))

        rows = self.get_rows(projects, request)
        url = self.get_download_url(request)
        response.context_data = {
            **response.context_data,
            **rows,
            **url
        }
        return response

    def get_rows(self, projects, request):
        ctxt = {}
        project_ids = {}
        for project in projects:
            project_ids.update({project.id: project})

        filtered_subsidy_period = None
        if self.subsidy_period_filter_param in request.GET:
            filtered_subsidy_period = int(request.GET[
                self.subsidy_period_filter_param
            ])
        else:
            current = SubsidyPeriod.get_current()
            if current:
                filtered_subsidy_period = current.id

        query = {
            # We need this count just to force the group_by
            'grouping': Count('project_id'),
            'members_h': Subquery(
                User.objects
                    .filter(gender='MALE')
                    .get_num_members_for_project(OuterRef('project_id')),
                output_field=IntegerField()
            ),
            'members_d': Subquery(
                User.objects
                    .filter(gender='FEMALE')
                    .get_num_members_for_project(OuterRef('project_id')),
                output_field=IntegerField()
            ),
            'members_total': Subquery(
                User.objects
                    .filter()
                    .get_num_members_for_project(OuterRef('project_id')),
                output_field=IntegerField()
            ),
            'creacio_hores': Sum(
                'stage_sessions__hours',
                filter=Q(stage_type=11)
            ),
            'creacio_certificat': Count(
                'scanned_certificate',
                filter=(
                    Q(stage_type=11) &
                    Q(scanned_certificate__isnull=False) &
                    ~Q(scanned_certificate__exact='')
                )
            ),
            'consolidacio_hores': Sum(
                'stage_sessions__hours',
                filter=Q(stage_type=12)
            ),
            'consolidacio_certificat': Count(
                'scanned_certificate',
                filter=(
                    Q(stage_type=12) &
                    Q(scanned_certificate__isnull=False) &
                    ~Q(scanned_certificate__exact='')
                )
            ),
            'incubation_hores': Sum(
                'stage_sessions__hours',
                filter=Q(stage_type=9)
            ),
            'incubation_certificat': Count(
                'scanned_certificate',
                filter=(
                    Q(stage_type=9) &
                    Q(scanned_certificate__isnull=False) &
                    ~Q(scanned_certificate__exact='')
                )
            ),
        }
        qs_project_stages = ProjectStage.objects.filter(
            project_id__in=project_ids
        )
        if filtered_subsidy_period:
            qs_project_stages = qs_project_stages.filter(
                subsidy_period=filtered_subsidy_period
            )

        # Annotate adds columns to each row with the sum or calculations of
        qs_project_stages = (
            qs_project_stages
                .values('project_id')
                .annotate(**query)
        )
        # Order_by fucks up the group by
        qs_project_stages = qs_project_stages.order_by()

        ctxt["rows"] = qs_project_stages

        for key, row in enumerate(ctxt['rows']):
            ctxt['rows'][key]['project'] = project_ids[row['project_id']]
            ctxt['rows'][key]['employment_insertions'] = len(
                row['project'].employment_insertions.filter(
                    subsidy_period=filtered_subsidy_period
                )
            )
            ctxt['rows'][key]['constituted'] = 0
            project_subsidy = ctxt['rows'][key]['project'].subsidy_period
            if project_subsidy:
                project_subsidy = project_subsidy.id
                if (
                    ctxt['rows'][key]['project'].constitution_date
                    and project_subsidy == filtered_subsidy_period
                    and ctxt['rows'][key]['project'].cif
                ):
                    ctxt['rows'][key]['constituted'] = 1

        ctxt["rows"] = self.sort_rows(ctxt["rows"])

        # Normally it should be easier to call aggregate to have the totals,
        # but given how complex is it to combine
        # with the ORM filters, I opted for filling the values like this.
        # The original version was: qs.aggregate(**query)
        totals = dict(
            total_members_h=0,
            total_members_d=0,
            total_members_total=0,
            total_creacio_hores=0,
            total_creacio_certificat=0,
            total_consolidacio_hores=0,
            total_consolidacio_certificat=0,
            total_incubation_hores=0,
            total_incubation_certificat=0,
            total_employment_insertions=0,
            total_constitutions=0,
            show_incubation=False
        )
        for row in ctxt['rows']:
            totals['total_members_h'] += (
                row['members_h'] if row['members_h'] else 0
            )
            totals['total_members_d'] += (
                row['members_d'] if row['members_d'] else 0
            )
            totals['total_members_total'] += (
                row['members_total'] if row['members_total'] else 0
            )
            totals['total_creacio_hores'] += (
                row['creacio_hores'] if row['creacio_hores'] else 0
            )
            totals['total_creacio_certificat'] += (
                1 if row['creacio_certificat'] else 0
            )
            totals['total_consolidacio_hores'] += (
                row['consolidacio_hores'] if row['consolidacio_hores'] else 0
            )
            totals['total_consolidacio_certificat'] += (
                1 if row['consolidacio_certificat'] else 0
            )
            totals['total_incubation_hores'] += (
                row['incubation_hores'] if row['incubation_hores'] else 0
            )
            totals['total_incubation_certificat'] += (
                1 if row['incubation_certificat'] else 0
            )
            totals['total_employment_insertions'] += row[
                'employment_insertions'
            ]
            totals['total_constitutions'] += row['constituted']

        if (totals['total_incubation_certificat'] > 0
                or totals['total_incubation_hores'] > 0):
            totals['show_incubation'] = True

        ctxt['totals'] = totals

        return ctxt

    @staticmethod
    def sort_rows(rows: list) -> list:
        """
        The user needs to apply ordering at the resulting report but we cannot
        rely on the ORM to do it, because if you apply an ordering to the
        annotated queryset that return the rows it stops working.

        Also, the ordering that we want is related to Project but the queryset
        with the annotated results queries ProjectStage.

        With this method, in order to modify the ordering you need to change
        the ordered_key composition.
        If instead of:
        ordered_key = f"{project.project_status}_{project.id}"

        You want to sort them by sector, you'll do:
        ordered_key = project.sector

        :param rows: All the project stages from the annotation queryset.
        :return: List with the same items but sorted differently
        """
        ordered_rows = {}
        for key, row in enumerate(rows):
            ordered_key = f"{row['project'].project_status}_{row['project'].id}"
            ordered_rows.update({
                ordered_key: rows[key]
            })
        rows = []
        for row in sorted(ordered_rows):
            rows.append(ordered_rows[row])
        return rows

    def get_download_url(self, request):
        querystring = urlencode(request.GET)
        url = reverse('admin:followup-spreadsheet')
        url = f"{url}?{querystring}"
        return {'spreadsheet_url': url}

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False

    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path(
                'spreadsheet/',
                self.admin_site.admin_view(self.download_spreadsheet),
                name='followup-spreadsheet',
            ),
        ]
        return custom_urls + urls

    def download_spreadsheet(self, request):
        # Getting queryset with filters applied
        try:
            qs = super().get_changelist_instance(request).queryset
        except (AttributeError, KeyError) as e:
            return request

        # Els filtres fan joins, necessitem el Count per forçar un group by.
        projects = qs.annotate(Count('id'))

        rows = self.get_rows(projects, request)
        sheet = FollowUpSpreadsheet(rows['rows'], rows['totals'])
        return sheet.export_seguiment_acompanyaments()


@admin.register(ProjectsFollowUpService)
class ProjectsFollowUpServicesAdmin(ProjectsFollowUpAdmin):
    change_list_template = 'admin/projects_follow_up_services.html'


class FollowUpSpreadsheet:
    def __init__(self, raw_rows, raw_totals):
        self.row_number = 1
        self.raw_rows = raw_rows
        self.raw_totals = raw_totals
        self.workbook = Workbook()
        self.worksheet = self.workbook.active

    def return_document(self, name):
        """ Attention: non-ascii characters in the name will cause
        an encoding error with gunicorn.
        Haven't tried it with a proxy under apache, in theory should
        work."""
        t = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response = HttpResponse(
            content_type=t,
        )
        date = datetime.now().strftime('%Y-%m-%d')
        response['Content-Disposition'] = (
            f'attachment; filename={date}-{name}.xlsx'
        )
        self.workbook.save(response)
        return response

    def create_columns(self, columns):
        """ create_columns

        Expects an iterable containing tuples with the name and the
        width of each column, like this:
        columns = [
            ("First", 40),
            ("Second", 70),
        ]
        """
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            column_letter = get_column_letter(col_num)
            column_dimensions = self.worksheet.column_dimensions[column_letter]
            column_dimensions.font = Font(name="ttf-opensans", size=9)
            column_dimensions.width = column_width
            cell.font = Font(bold=True, name="ttf-opensans", size=9)
            cell.border = Border(
                bottom=Side(border_style="thin", color='000000')
            )
            cell.value = str(column_title)

    def fill_row_data(self, row):
        """ fill_row_data

        Populates the columns of a given row with each of the values.
        Expects an iterable with the data for each row:
        row = [
            "first value",
            "second value",
        ]

        Optionally, values can be a tuple to mark the cell as error.
        That will fill the cell with red.
        row = [
            "first value",
            ("second value", True),
        ]
        """
        for col_num, cell_value in enumerate(row, 1):
            cell = self.worksheet.cell(row=self.row_number, column=col_num)
            if isinstance(cell_value, tuple):
                error_mark = cell_value[1]
                if error_mark:
                    cell.fill = PatternFill(
                        start_color='FFFF0000',
                        end_color='FFFF0000',
                        fill_type='solid'
                    )
                cell_value = cell_value[0]
            cell.value = (cell_value
                          if isinstance(cell_value, int) else str(cell_value))

    def export_seguiment_acompanyaments(self):
        """ Each function here called handles the creation of one of the
        worksheets."""
        self.export_seguiment_acompanyaments_sheet1()

        return self.return_document("seguiment_acompanyaments")

    def export_seguiment_acompanyaments_sheet1(self):
        # Tutorial: https://djangotricks.blogspot.com/2019/02/how-to-export-data-to-xlsx-files.html
        # Docs: https://openpyxl.readthedocs.io/en/stable/tutorial.html#create-a-workbook
        self.worksheet.title = "Acompanyaments"

        columns = [
            ("ID", 10),
            ("Seguiment", 35),
            ("Ateneu/Cercle", 20),
            ("Nom", 50),
            ("Servei", 30),
            ("Tutoritza", 20),
            ("Membres H", 15),
            ("Membres D", 15),
            ("Total sòcies", 15),
            ("Sector", 20),
            ("Descripció", 60),
            ("Estat", 20),
            ("Territori", 20),
            ("Creació H.", 15),
            ("Creació cert.", 15),
            ("Consolidació H.", 15),
            ("Consolidació cert.", 15),
            ("Incubació H.", 15),
            ("Incubació cert.", 15),
            ("Insercions previstes", 15),
            ("Insercions justificades", 15),
            ("Constitució", 15),
            ("Altres", 60),
        ]
        self.create_columns(columns)
        self.export_seguiment_acompanyaments_rows()

    def export_seguiment_acompanyaments_rows(self):
        self.row_number = 1
        for raw_row in self.raw_rows:
            self.row_number += 1
            print(raw_row)
            stage_responsible = raw_row['project'].last_stage_responsible
            follow_up_situation = \
                raw_row['project'].get_follow_up_situation_display()
            row = [
                raw_row['project'].id,
                follow_up_situation if follow_up_situation else '',
                raw_row["project"].last_stage_circle,
                raw_row['project'].name,
                (raw_row['project'].services_list
                    if raw_row['project'].services_list else ''),
                stage_responsible if stage_responsible else '',
                raw_row['members_h'],
                raw_row['members_d'],
                raw_row['members_total'],
                raw_row['project'].get_sector_display(),
                (raw_row['project'].description
                    if raw_row['project'].description else ''),
                raw_row['project'].get_project_status_display(),
                (raw_row['project'].full_town_district
                    if raw_row['project'].full_town_district else ''),
                (raw_row['creacio_hores']
                    if raw_row['creacio_hores'] else 0),
                1 if raw_row['creacio_certificat'] > 0 else 0,
                (raw_row['consolidacio_hores']
                    if raw_row['consolidacio_hores'] else 0),
                1 if raw_row['consolidacio_certificat'] > 0 else 0,
                (raw_row['incubation_hores']
                    if raw_row['incubation_hores'] else 0),
                1 if raw_row['incubation_certificat'] > 0 else 0,
                raw_row['project'].employment_estimation,
                len(raw_row['project'].employment_insertions.all()),
                (raw_row['project'].constitution_date
                    if raw_row['project'].constitution_date else ''),
                raw_row['project'].other if raw_row['project'].other else '',
            ]
            self.fill_row_data(row)


class ConstitutionDateFilter(admin.SimpleListFilter):
    title = "Amb data de constitució dins la convocatòria…"

    parameter_name = 'constitution_subsidy'

    def lookups(self, request, model_admin):
        qs = SubsidyPeriod.objects.all()
        qs.order_by('name')
        return list(qs.values_list('id', 'name'))

    def queryset(self, request, queryset):
        value = self.value()
        if value:
            period = SubsidyPeriod.objects.get(id=value)
            return queryset.filter(constitution_date__range=(
                period.date_start, period.date_end)
            )
        return queryset


@admin.register(ProjectsConstituted)
class ProjectsConstitutedAdmin(admin.ModelAdmin):
    """
    Inspired in: https://medium.com/@hakibenita/how-to-turn-django-admin-into-a-lightweight-dashboard-a0e0bbf609ad
    """
    class Media:
        js = ('js/grappellihacks.js',)
        css = {
            'all': ('styles/grappellihacks.css',)
        }

    change_list_template = 'admin/projects_constituted.html'
    list_filter = (ConstitutionDateFilter, )
    show_full_result_count = False
    list_display = ('name', )
    list_per_page = 99999

    def changelist_view(self, request, extra_context=None):
        response = super().changelist_view(
            request,
            extra_context=extra_context,
        )

        try:
            qs = response.context_data['cl'].queryset
        except (AttributeError, KeyError):
            return response

        query = {
            'members_h': Count(
                'stages__involved_partners',
                filter=Q(stages__involved_partners__gender='MALE'),
                distinct=True
            ),
            'members_d': Count(
                'stages__involved_partners',
                filter=Q(stages__involved_partners__gender='FEMALE'),
                distinct=True
            ),
            'members_a': Count(
                'stages__involved_partners',
                filter=~Q(stages__involved_partners__gender='FEMALE')
                       & ~Q(stages__involved_partners__gender='MALE'),
                distinct=True
            ),
        }

        # Annotate adds columns to each row with the sum or calculations of the row:
        response.context_data['rows'] = list(
            qs.filter(
                cif__isnull=False, constitution_date__isnull=False
            ).annotate(**query)
        )

        # Normally it should be easier to call aggregate to have the totals,
        # but given how complex is it to combine
        # with the ORM filters, I opted for filling the values like this.
        # The original version was: qs.aggregate(**query)
        totals = dict(
            total_members_h=0,
            total_members_d=0,
            total_members_a=0,
        )
        for row in response.context_data['rows']:
            totals['total_members_h'] += row.members_h if row.members_h else 0
            totals['total_members_d'] += row.members_d if row.members_d else 0
            totals['total_members_a'] += row.members_a if row.members_a else 0

        response.context_data['totals'] = totals

        return response

    # def has_change_permission(self, request, obj=None):
    #     return False

    def has_delete_permission(self, request, obj=None):
        return False

    def has_add_permission(self, request):
        return False


@admin.register(ProjectsConstitutedService)
class ProjectsConstitutedServiceAdmin(ProjectsConstitutedAdmin):
    change_list_template = 'admin/projects_constituted_service.html'
