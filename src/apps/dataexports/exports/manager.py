import json

from django.http import HttpResponseNotFound, HttpResponse
from openpyxl import Workbook
from datetime import datetime
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, PatternFill
from django.conf import settings


class ExportManager:
    correlations = None

    def __init__(self, export_obj):
        self.error_message = set()
        self.ignore_errors = export_obj.ignore_errors
        self.subsidy_period = export_obj.subsidy_period
        self.subsidy_period_range = (
            export_obj.subsidy_period.range
        )

    def return_404(self, message=""):
        """When the exported data has to fit a specific format, there
        are many cases in which we need to stop the generation and tell
        the user that something needs to be fixed.
        This will show a blank page with the message.
        """
        if message:
            self.error_message.add(message)
        message = "<h1>Error al generar el document</h1>" + " ".join(
            self.error_message)
        return HttpResponseNotFound(message)

    def import_correlations(self, file_path):
        try:
            file_object = open(file_path, 'r')
            self.correlations = json.load(file_object)
        except FileNotFoundError:
            print(file_path + " not found. ")

    def get_correlation(self, correlated_field, original_data, subsidy_period=2019):
        """When exporting data, we might need to make the exported data
         fit specific requirements. For example, we store the field
         'axis' as 'A', 'B', but the strings we actually need to
         show are:
         'A) Diagnosi i visibilització', 'B) Creació i desenvolupament'

         We have these correlations in a json file and loaded at self.correlations.

        This function is a wrapper to get those.
        """
        try:
            new_data = self.correlations[correlated_field][original_data]
        except KeyError:
            self.error_message.add(
                "<p>El document no s'ha pogut generar perquè s'ha intentat aplicar aquesta correlació:</p"
                "<ul><li>Convocatòria: {}</li><li>Camp: {}</li><li>Dada original: {}</li></ul>"
                "<p>Però no s'ha trobat.</p>".format(subsidy_period, correlated_field, original_data))
            return None
        return new_data


class ExcelExportManager(ExportManager):
    def __init__(self, export_obj):
        super().__init__(export_obj)
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.row_number = 1
        self.correlations = dict()

        self.import_correlations(
            settings.BASE_DIR
            + "/apps/dataexports/fixtures/correlations_2019.json"
        )

    def return_document(self, name):
        """ Attention: non-ascii characters in the name will cause
        an encoding error with gunicorn.
        Haven't tried it with a proxy under apache, in theory should
        work."""
        if len(self.error_message) > 0 and self.ignore_errors is False:
            return self.return_404()

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename={date}-{name}.xlsx'.format(
            date=datetime.now().strftime('%Y-%m-%d'), name=name,
        )
        self.workbook.save(response)
        return response

    def create_columns(self, columns):
        """ create_columns

        Expects an iterable containing tuples with the name and the
        width of each column, like this:
        columns = [
            ("First", 40),
            ("Second", 70),
        ]
        """
        for col_num, (column_title, column_width) in enumerate(columns, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            column_letter = get_column_letter(col_num)
            column_dimensions = self.worksheet.column_dimensions[column_letter]
            column_dimensions.font = Font(name="ttf-opensans", size=9)
            column_dimensions.width = column_width
            cell.font = Font(bold=True, name="ttf-opensans", size=9)
            cell.border = Border(bottom=Side(border_style="thin", color='000000'))
            cell.value = str(column_title)

    def fill_row_data(self, row):
        """ fill_row_data

        Populates the columns of a given row with each of the values.
        Expects an iterable with the data for each row:
        row = [
            "first value",
            "second value",
        ]

        Optionally, values can be a tuple to mark the cell as error.
        That will fill the cell with red.
        row = [
            "first value",
            ("second value", True),
        ]
        """
        for col_num, cell_value in enumerate(row, start=1):
            cell = self.worksheet.cell(row=self.row_number, column=col_num)
            if isinstance(cell_value, tuple):
                error_mark = cell_value[1]
                if error_mark:
                    cell.fill = PatternFill(start_color='FFFF0000', end_color='FFFF0000', fill_type='solid')
                cell_value = cell_value[0]
            cell.value = cell_value

    def fill_row_from_factory(self, row_factory):
        self.row_number += 1
        self.fill_row_data(row_factory.get_columns())
        if row_factory.get_format_method():
            getattr(self, row_factory.get_format_method()[0])(
                row_factory.get_format_method()[1]
            )

    def format_row(self, row_num, prop_name, obj):
        """
        Applies the given format to the given property to each of the cells of
         the row.

        :param row_num: int, row nº to modify.
        :param prop_name: str, name of the property, ex: "fill"
        :param obj: one of the openpyxl.styles objects
        :return:
        """
        for col in range(1, self.worksheet.max_column + 1):
            self.format_cell(col, row_num, prop_name, obj)

    def format_cell(self, col_num, row_num, prop_name, obj):
        cell = self.worksheet.cell(column=col_num, row=row_num)
        setattr(cell, prop_name, obj)

    def format_cell_bold(self, col_num, row_num: int = None):
        if not row_num:
            row_num = self.row_number
        format_obj = Font(bold=True, name="ttf-opensans", size=9)
        self.format_cell(col_num, row_num, "font", format_obj)

    def format_row_header(self, row_num: int = None):
        if not row_num:
            row_num = self.row_number
        format_obj = Font(bold=True, name="ttf-opensans", size=9)
        self.format_row(
            row_num,
            "font",
            format_obj
        )
        format_obj = Border(bottom=Side(border_style="thin", color='000000'))
        self.format_row(
            row_num,
            "border",
            format_obj
        )

    def format_cell_default_font(self, col_num, row_num):
        format_obj = Font(name="ttf-opensans", size=9)
        self.format_cell(
            col_num,
            row_num,
            "font",
            format_obj
        )

    def set_cell_value(self, col_num, row_num, value):
        cell = self.worksheet.cell(row=row_num, column=col_num)
        cell.value = value
